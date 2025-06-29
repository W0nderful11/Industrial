import logging
import json
import re
from typing import Optional, Dict, Any, List, Tuple
import openpyxl
from openpyxl.utils import get_column_letter

# Импортируем общие функции и константы
from .utils import KNOWN_ERROR_CODES, filter_cell
# Импортируем ИИ функции для полного анализа
from services.telegram.ai.ai import analyze_image_via_ai

# --- Функция для очистки строк с пробелами (из предоставленного кода) ---
def clean_spaced_string(s):
    return re.sub(r"\s+", "", s)


class PhotoLogAnalyzer:
    """Анализатор логов для фотографий - использует только ИИ"""
    
    def __init__(self, lang):
        self.panic_sheet = None
        self.lang = lang
        # logging.info(f"Initializing PhotoLogAnalyzer for language: {lang}")

        # Load panic_codes.xlsx
        try:
            panic_workbook = openpyxl.load_workbook("./data/panic_codes.xlsx")
            try:
                self.panic_sheet = panic_workbook[lang]
                # logging.info(f"Loaded sheet '{lang}' from panic_codes.xlsx")
            except KeyError:
                # Fallback to active sheet if language sheet not found
                self.panic_sheet = panic_workbook.active
                logging.warning(
                    f"Sheet '{lang}' not found in panic_codes.xlsx, using active sheet: '{self.panic_sheet.title}'")
        except FileNotFoundError:
            logging.error("File ./data/panic_codes.xlsx not found!")
        except Exception as e:
            logging.error(f"Error loading panic_codes.xlsx: {e}")

    def _find_solution_by_code(self, sheet, product_key, error_code_to_find):
        """
        Searches for an exact match of error_code_to_find in column 'A' of the sheet
        and returns the solution for the given product_key.
        Returns tuple (solutions, links) or (None, None) if not found.
        """
        if not sheet:
            logging.warning(f"_find_solution_by_code: Sheet is None, cannot search.")
            return None, None
        if not product_key or product_key.lower() == "неизвестно":
            logging.warning(f"_find_solution_by_code: Invalid product_key ('{product_key}'), cannot search.")
            return None, None
        if not error_code_to_find:
            logging.warning(f"_find_solution_by_code: error_code_to_find is empty or None, cannot search.")
            return None, None

        model_column_index = None
        try:
            # Assuming model headers are in the second row
            header_row = sheet[2]
        except IndexError:
            logging.error(f"Could not read header row (2) in sheet '{sheet.title}'")
            return None, None

        # Find the column index for our model identifier
        product_key_normalized = str(product_key).lower().replace(" ", "")
        for cell in header_row:
            if cell.value:
                cell_value_normalized = str(cell.value).lower().replace(" ", "")
                if cell_value_normalized == product_key_normalized:
                    model_column_index = cell.column
                    logging.info(
                        f"Found column {get_column_letter(model_column_index)} for model '{product_key}' in sheet '{sheet.title}'.")
                    break

        if model_column_index is None:
            logging.warning(
                f"Column for model '{product_key}' (normalized key: '{product_key_normalized}') not found in sheet '{sheet.title}'.")
            return None, None

        # Search for the error code in column 'A' (starting from row 3)
        error_code_lower = error_code_to_find.lower()

        for row_index in range(3, sheet.max_row + 1):
            error_code_cell = sheet.cell(row=row_index, column=1).value
            if not error_code_cell:
                continue

            error_code_in_sheet = str(error_code_cell).strip().lower()
            # Case-insensitive comparison
            if error_code_in_sheet == error_code_lower:
                solution_cell = sheet.cell(row=row_index, column=model_column_index).value
                solution_text = str(solution_cell or "").strip()
                if solution_text:
                    # Используем filter_cell для разбора решения
                    solutions, links = filter_cell(solution_text)
                    return solutions, links
                else:
                    logging.warning(
                        f"Found code '{error_code_to_find}' in row {row_index}, but solution cell for model '{product_key}' is empty.")
                    return None, None

        return None, None


# --- Класс PhotoAnalyzer для анализа фотографий с помощью ИИ ---
class PhotoAnalyzer:
    """
    Класс для анализа фотографий логов iOS с помощью ИИ.
    Использует OpenAI GPT-4o для полного анализа изображений.
    Интегрирован с обработчиком поиска ошибок в Excel и ИИ промпте.
    """
    
    def __init__(self, lang: str, file_path: str, username: Optional[str] = None):
        self.lang = lang
        self.file_path = file_path
        self.username = username
        self.log_data = {}
        self.log_analyzer = PhotoLogAnalyzer(lang)
        # logging.info(f"PhotoAnalyzer: Инициализирован для анализа {file_path}")
        
    def get_model(self):
        """Возвращает модель устройства из данных ИИ анализа, находя человекочитаемое имя в Excel."""
        from services.telegram.schemas.analyzer import ModelPhone
        
        product_identifier = self.log_data.get('product') # e.g., iPhone14,4
        ios_version = self.log_data.get('os_version', "Неизвестно")
        crash_key = self.log_data.get('crash_reporter_key')

        if not product_identifier:
            return ModelPhone(
                model="Неизвестно",
                version="Неизвестно",
                crash_reporter_key=crash_key,
                ios_version=ios_version
            )

        sheet = self.log_analyzer.panic_sheet
        if not sheet:
            logging.warning("PhotoAnalyzer: panic_sheet is not loaded. Cannot resolve model name.")
            return ModelPhone(
                model=product_identifier,
                version=product_identifier,
                crash_reporter_key=crash_key,
                ios_version=ios_version
            )

        product_normalized = product_identifier.lower().replace(" ", "")

        try:
            # Ищем идентификатор во второй строке
            for cell_obj in sheet[2]: 
                if cell_obj.value and isinstance(cell_obj.value, str):
                    cell_value_normalized = cell_obj.value.lower().replace(" ", "")
                    if cell_value_normalized == product_normalized:
                        # Нашли совпадение, берем имя из первой строки того же столбца
                        model_name_cell = sheet.cell(row=1, column=cell_obj.column)
                        model_name = model_name_cell.value if model_name_cell and model_name_cell.value else product_identifier
                        
                        return ModelPhone(
                            model=model_name, # e.g., iPhone 13 mini
                            version=product_identifier, # e.g., iPhone14,4
                            crash_reporter_key=crash_key,
                            ios_version=ios_version
                        )
        except IndexError:
             logging.error(f"Could not read header rows in sheet '{sheet.title}' for model lookup.")
             # Fallback to returning the identifier
        
        # Если не нашли в Excel
        logging.warning(f"PhotoAnalyzer: Product '{product_identifier}' not found in Excel headers.")
        return ModelPhone(
            model=product_identifier,
            version=product_identifier,
            crash_reporter_key=crash_key,
            ios_version=ios_version
        )
        
    async def find_error_solutions(self, debug: bool = False):
        from services.telegram.schemas.analyzer import SolutionAboutError
        
        # Инициализация переменных для хранения решений
        mini_sols, mini_lnks = [], []
        full_sols, full_lnks = [], []
        determined_error_code = None
        panic_string_from_ai = ""

        try:
            ai_result = await analyze_image_via_ai(self.file_path, KNOWN_ERROR_CODES)
            
            if ai_result and isinstance(ai_result, dict):
                crash_key = ai_result.get('crash_reporter_key')
                if crash_key:
                    crash_key = crash_key.lower()

                self.log_data = {
                    'product': ai_result.get('product'),
                    'os_version': ai_result.get('os_version'),
                    'timestamp': ai_result.get('timestamp'),
                    'panic_string': ai_result.get('panic_string', ''),
                    'crash_reporter_key': crash_key
                }
                determined_error_code = ai_result.get('error_code')
                panic_string_from_ai = self.log_data.get('panic_string', '')

                if determined_error_code:
                    product_key = self.log_data.get('product')
                    if product_key and self.log_analyzer.panic_sheet:
                        # 1. Ищем "полное" решение
                        f_sols, f_lnks = self.log_analyzer._find_solution_by_code(
                            self.log_analyzer.panic_sheet, product_key, determined_error_code
                        )
                        if f_sols: full_sols.extend(f_sols)
                        if f_lnks: full_lnks.extend(f_lnks)
                        
                        # 2. Ищем "мини" решение
                        mini_error_code_target = determined_error_code + " mini"
                        m_sols, m_lnks = self.log_analyzer._find_solution_by_code(
                            self.log_analyzer.panic_sheet, product_key, mini_error_code_target
                        )
                        if m_sols: mini_sols.extend(m_sols)
                        if m_lnks: mini_lnks.extend(m_lnks)
            
            elif ai_result == 'RATE_LIMIT_EXHAUSTED':
                # Обработка ошибки лимита запросов (возвращаем сообщение об ошибке)
                return SolutionAboutError(
                    descriptions=[], links=[], date_of_failure="", is_full=False, error_code=None,
                    panic_string='AI analysis unavailable: Rate limit exceeded',
                    extracted_error_text_for_admin='AI analysis unavailable: Rate limit exceeded'
                )
            elif ai_result == 'TIMEOUT_EXHAUSTED':
                # Обработка ошибки таймаута (возвращаем сообщение об ошибке)
                return SolutionAboutError(
                    descriptions=[], links=[], date_of_failure="", is_full=False, error_code=None,
                    panic_string='AI analysis unavailable: Timeout exceeded',
                    extracted_error_text_for_admin='AI analysis unavailable: Timeout exceeded'
                )
            else: # ai_result is None или другая ошибка
                return SolutionAboutError(
                    descriptions=[], links=[], date_of_failure=self.log_data.get('timestamp', ''), 
                    is_full=False, error_code=None, 
                    panic_string="Could not analyze image or extract error code.",
                    extracted_error_text_for_admin="Could not analyze image or extract error code."
                )

        except Exception as e:
            logging.error(f"PhotoAnalyzer: Exception during AI analysis or Excel lookup: {e}", exc_info=True)
            return SolutionAboutError(
                descriptions=[], links=[], date_of_failure=self.log_data.get('timestamp', ''), 
                is_full=False, error_code=None, panic_string=f"Error during photo analysis: {e}",
                extracted_error_text_for_admin=f"Error during photo analysis: {e}"
            )

        # Логика формирования SolutionAboutError, аналогичная BaseAnalyzer
        user_facing_descriptions = []
        user_facing_links = []
        is_mini_shown = False
        has_full_available = False
        final_full_descriptions_for_storage = None
        final_full_links_for_storage = None

        if mini_sols or mini_lnks:
            user_facing_descriptions.extend(mini_sols)
            user_facing_links.extend(mini_lnks)
            is_mini_shown = True
            if full_sols or full_lnks:
                has_full_available = True
                final_full_descriptions_for_storage = full_sols
                final_full_links_for_storage = full_lnks
        elif full_sols or full_lnks:
            user_facing_descriptions.extend(full_sols)
            user_facing_links.extend(full_lnks)
        
        legacy_is_full_flag = bool(user_facing_descriptions or user_facing_links)
        
        # panic_string_from_ai уже должен быть обработан (до 'slide') на уровне AI промпта
        # extracted_error_text_for_admin будет panic_string_from_ai
        
        return SolutionAboutError(
            descriptions=user_facing_descriptions,
            links=user_facing_links,
            date_of_failure=self.log_data.get('timestamp', ''),
            is_full=legacy_is_full_flag,
            error_code=determined_error_code,
            panic_string=panic_string_from_ai, 
            extracted_error_text_for_admin=panic_string_from_ai, 
            is_mini_response_shown=is_mini_shown,
            has_full_solution_available=has_full_available,
            full_descriptions=final_full_descriptions_for_storage,
            full_links=final_full_links_for_storage
        ) 