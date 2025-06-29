import openpyxl
from typing import List, Tuple, Optional
from config import PANIC_CODES_EXCEL_PATH, DEFAULT_SHEET_NAME_FOR_CODES


def load_error_codes_from_excel() -> List[str]:
    """Загружает известные коды ошибок из столбца А указанного листа Excel."""
    workbook = openpyxl.load_workbook(PANIC_CODES_EXCEL_PATH)
    worksheet = workbook[
        DEFAULT_SHEET_NAME_FOR_CODES if DEFAULT_SHEET_NAME_FOR_CODES in workbook.sheetnames else workbook.active
    ]

    codes = []
    for code in worksheet.iter_rows(min_row=3, max_col=1, values_only=True):
        if code and code[0]:
            original_code = str(code[0]).strip()
            codes.append(original_code)
            
            # Добавляем вариант без экранирования для кодов с \/
            if '\\/' in original_code:
                unescaped_code = original_code.replace('\\/', '/')
                if unescaped_code not in codes:
                    codes.append(unescaped_code)

    return codes


def reload_known_error_codes() -> List[str]:
    """
    Перезагружает список известных кодов ошибок из Excel файла.
    Используется после замены panic_codes.xlsx файла.
    
    Returns:
        Обновленный список кодов ошибок
    """
    global KNOWN_ERROR_CODES
    KNOWN_ERROR_CODES = load_error_codes_from_excel()
    return KNOWN_ERROR_CODES


def filter_cell(text: Optional[str]) -> Tuple[List[str], List[str]]:
    """
    Разбирает текст ячейки Excel на решения и ссылки.
    Возвращает кортеж (solutions, links).
    """
    solutions = []
    links = []
    if text:
        for value in text.split(";"):
            if (value := value.strip()).startswith("http"):
                links.append(value)
            elif value:
                solutions.append(value)
    return solutions, links


# Загружаем коды ошибок один раз при импорте модуля
KNOWN_ERROR_CODES = load_error_codes_from_excel() 