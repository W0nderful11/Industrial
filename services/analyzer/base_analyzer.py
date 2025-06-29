import io
import re
import asyncio
from typing import Dict, List, Tuple, Optional
import os # Импортируем os для доступа к переменным окружения
import openai # Импортируем openai для создания клиента

import openpyxl
from openpyxl.workbook import Workbook
from PIL import Image
from openpyxl.utils import get_column_letter
import json
import logging
from datetime import datetime

from config import PANIC_CODES_EXCEL_PATH
from services.telegram.schemas.analyzer import ModelPhone, SolutionAboutError
from services.telegram.ai.ai import get_ai_error_code_suggestion
from .utils import filter_cell, KNOWN_ERROR_CODES


class BaseAnalyzer:
    def __init__(self, lang: str, path: Optional[str] = None, username: Optional[str] = None):
        self.lang = lang
        self.path = path
        self.username = username
        self.log = ""
        self.log_dict: Dict = {}
        self.sheet = None
        self._images = {}

        # Определяем путь к Excel файлу относительно корня проекта
        if os.path.exists("./data/panic_codes.xlsx"):
            excel_file_path = "./data/panic_codes.xlsx"
        elif os.path.exists("../data/panic_codes.xlsx"):
            excel_file_path = "../data/panic_codes.xlsx"
        else:
            # Используем константу из config.py если доступна
            try:
                excel_file_path = PANIC_CODES_EXCEL_PATH
            except NameError:
                excel_file_path = "./data/panic_codes.xlsx"
        
        try:
            workbook: Workbook = openpyxl.load_workbook(excel_file_path)
            self.sheet = workbook[lang]
        except FileNotFoundError:
            logging.error(f"ERROR: Excel file not found at {excel_file_path}")
            self.sheet = None
        except KeyError:
            logging.error(f"ERROR: Sheet '{lang}' not found in {excel_file_path}")
            self.sheet = None

        if path:
            self.load_and_parse_file()

    def load_and_parse_file(self) -> None:
        raise NotImplementedError

    def read_images(self) -> None:
        if self.sheet:
            try:
                # Пытаемся получить изображения из листа
                sheet_images = getattr(self.sheet, '_images', [])
                for image in sheet_images:
                    if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                        row = image.anchor._from.row + 1
                        col_letter = get_column_letter(image.anchor._from.col + 1)
                        self._images[f'{col_letter}{row}'] = image._data
            except Exception as e:
                logging.warning(f"Не удалось загрузить изображения из Excel: {e}")
                self._images = {}
        else:
            logging.warning("self.sheet не инициализирован в read_images.")
            self._images = {}

    def get_image(self, cell: str) -> Image.Image:
        if not self._images:
            self.read_images()

        if cell not in self._images:
            self.read_images()
            if cell not in self._images:
                raise ValueError(f"Cell {cell} doesn't contain an image, even after re-reading.")

        image_data_fn = self._images[cell]
        image_bytes = image_data_fn()
        image_stream = io.BytesIO(image_bytes)
        return Image.open(image_stream)

    def get_model(self) -> Optional[ModelPhone]:
        # Пытаемся получить crash_reporter_key из log_dict, проверяя оба стиля именования
        crash_key_from_log = self.log_dict.get('crash_reporter_key') or self.log_dict.get('crashReporterKey')
        if crash_key_from_log:
            crash_key_from_log = crash_key_from_log.lower()

        product_from_log = self.log_dict.get("product")
        os_version_from_log = self.log_dict.get('build') or self.log_dict.get('os_version')
        # productVersion может быть более специфичным, чем просто product (например, для идентификаторов моделей)
        product_version_from_log = self.log_dict.get("productVersion") or product_from_log

        if not product_from_log:
            # Если product не найден, все равно возвращаем ModelPhone с тем, что есть, особенно crash_reporter_key
            return ModelPhone(
                model="Неизвестно",
                version="Неизвестно",
                crash_reporter_key=crash_key_from_log,
                ios_version=os_version_from_log or "Неизвестно"
            )

        if self.sheet is None:
            # Если нет Excel-таблицы, возвращаем данные только из log_dict
            # Это важно, чтобы не терять crash_reporter_key
            logging.warning("Excel sheet is None in get_model. Returning data from log_dict only.")
            return ModelPhone(
                model=product_from_log,
                version=product_version_from_log,
                crash_reporter_key=crash_key_from_log,
                ios_version=os_version_from_log or "Неизвестно"
            )

        product_normalized = product_from_log.lower().replace(" ", "")

        for cell_obj in self.sheet[2]: # Версии продукта (например, iPhone10,1) во второй строке
            if isinstance(cell_obj.value, str):
                cell_value_as_string = cell_obj.value
                cell_value_normalized = cell_value_as_string.lower().replace(" ", "")
                if cell_value_normalized == product_normalized:
                    # Проверяем, что column не None
                    if cell_obj.column is not None:
                        model_name_cell = self.sheet.cell(row=1, column=cell_obj.column) # Имя модели (например, iPhone X) в первой строке
                        model_name = str(model_name_cell.value) if model_name_cell and model_name_cell.value else product_from_log
                        
                        return ModelPhone(
                            model=model_name,
                            version=cell_value_as_string, # Это версия продукта (iPhone10,1)
                            crash_reporter_key=crash_key_from_log,
                            ios_version=os_version_from_log or "Неизвестно"
                        )
        
        # Если product из лога не найден в заголовках Excel, возвращаем данные из log_dict
        logging.warning(f"Product '{product_from_log}' not found in Excel headers. Returning data from log_dict.")
        return ModelPhone(
            model=product_from_log,
            version=product_version_from_log,
            crash_reporter_key=crash_key_from_log,
            ios_version=os_version_from_log or "Неизвестно"
        )

    def _get_all_known_error_codes_from_excel(self, debug: bool = False) -> List[str]:
        """
        Возвращает список всех известных кодов ошибок.
        Использует обновленный KNOWN_ERROR_CODES из utils.py который обрабатывает экранированные символы.
        """
        if debug:
            pass
            # print(f"DEBUG: Using KNOWN_ERROR_CODES from utils.py. Total codes: {len(KNOWN_ERROR_CODES)}")
        
        # Фильтруем mini коды как в оригинальной логике
        filtered_codes = [code for code in KNOWN_ERROR_CODES if " mini" not in code.lower()]
        
        if debug:
            pass
            # print(f"DEBUG: After filtering mini codes: {len(filtered_codes)} codes")
        
        return filtered_codes

    async def _get_error_code_via_ai(self, extracted_error_text: str, debug: bool = False) -> Optional[str]:
        if not extracted_error_text:
            if debug:
                pass
                # print("DEBUG (AI): No extracted error text to send to AI.")
            return None

        all_known_codes = self._get_all_known_error_codes_from_excel(debug=debug)
        if not all_known_codes:
            if debug:
                pass
                # print("DEBUG (AI): No known error codes from Excel to provide to AI.")
            return None
        
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            logging.error("OPENAI_API_KEY не найден для _get_error_code_via_ai в BaseAnalyzer.")
            return None
        
        # Создаем клиент OpenAI здесь, т.к. этот метод вызывается для текстовых файлов
        client = openai.AsyncOpenAI(api_key=api_key)

        if debug:
            pass
            # print(f"DEBUG (AI): Sending to AI - Extracted Text: '{extracted_error_text[:200]}...'")

        ai_suggested_code = await get_ai_error_code_suggestion(
            client=client, # Передаем созданный клиент
            error_text=extracted_error_text,
            known_error_codes=all_known_codes
        )

        if debug:
            pass
            # print(f"DEBUG (AI): AI suggested error code: '{ai_suggested_code}'")
        if not ai_suggested_code:
            if debug:
                pass
                # print("DEBUG (AI): AI did not suggest an error code.")
            return None

        if ai_suggested_code and ai_suggested_code not in all_known_codes:
            if debug:
                pass
                # print(f"DEBUG (AI): Warning - AI suggested code '{ai_suggested_code}' is NOT in the Excel list of known codes. Treating as no match.")
            return None

        return ai_suggested_code

    def _get_solutions_for_excel_code(self, error_code_from_ai: str, model_column: Optional[int], debug: bool = False) -> Optional[Dict]:
        if self.sheet is None or not error_code_from_ai or model_column is None:
            if debug:
                pass
                # print(f"DEBUG (Excel Solution): Invalid parameters - sheet: {self.sheet is not None}, code: '{error_code_from_ai}', model_col: {model_column}")
            return None

        for index, row_data_tuple in enumerate(self.sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row_data_tuple or not row_data_tuple[0]:
                continue
                
            current_excel_code = str(row_data_tuple[0]).replace('"', '').strip()
            
            # Нормализуем оба кода для сравнения - убираем экранирование
            normalized_excel_code = current_excel_code.replace('\\/', '/').lower()
            normalized_ai_code = error_code_from_ai.replace('\\/', '/').lower()

            if debug:
                pass
                # print(f"DEBUG (Excel Solution): Comparing Excel '{current_excel_code}' (normalized: '{normalized_excel_code}') with AI '{error_code_from_ai}' (normalized: '{normalized_ai_code}')")

            if normalized_excel_code == normalized_ai_code:
                if debug:
                    pass
                    # print(f"DEBUG (Excel Solution): Found match for AI code '{error_code_from_ai}' at Excel row ~{index}")

                solutions, links = [], []
                image_path = None

                if 0 <= (model_column - 1) < len(row_data_tuple) and row_data_tuple[model_column - 1]:
                    solutions, links = filter_cell(str(row_data_tuple[model_column - 1]))
                else:
                    if debug:
                        pass
                        # print(f"DEBUG (Excel Solution): No solution for model_column {model_column} (index {model_column - 1}), trying fallback for '{error_code_from_ai}'")
                    for col_idx_0_based in range(1, self.sheet.max_column):
                        if col_idx_0_based == (model_column - 1):
                            continue
                        if col_idx_0_based < len(row_data_tuple) and row_data_tuple[col_idx_0_based]:
                            fallback_solutions, fallback_links = filter_cell(str(row_data_tuple[col_idx_0_based]))
                            if fallback_solutions or fallback_links:
                                solutions.extend(fallback_solutions)
                                links.extend(fallback_links)
                                if debug:
                                    pass
                                    # print(f"DEBUG (Excel Solution): Found fallback solution in column index {col_idx_0_based} (Excel col {get_column_letter(col_idx_0_based + 1)}) for '{error_code_from_ai}'")
                                try:
                                    cell_for_image = f'{get_column_letter(col_idx_0_based + 1)}{index}'
                                    image = self.get_image(cell_for_image)
                                    path = f'./{self.username or "user"}_{cell_for_image}.png'
                                    image.save(path)
                                    image_path = path
                                except ValueError:
                                    if debug:
                                        pass
                                        # print(f"DEBUG (Excel Solution Image Fallback): No image at {cell_for_image}")
                                except Exception as e_img:
                                    if debug:
                                        pass
                                        # print(f"DEBUG (Excel Solution Image Fallback): Error processing image from {cell_for_image}: {e_img}")
                                break

                if not image_path and (solutions or links):
                    try:
                        cell_main_image = f'{get_column_letter(model_column)}{index}'
                        image = self.get_image(cell_main_image)
                        path = f'./{self.username or "user"}_{cell_main_image}.png'
                        image.save(path)
                        image_path = path
                    except ValueError:
                        if debug:
                            pass
                            # print(f"DEBUG (Excel Solution Image Main): No image at {cell_main_image}")
                        pass
                    except Exception as e_img_main:
                        if debug:
                            pass
                            # print(f"DEBUG (Excel Solution Image Main): Error processing image from {cell_main_image}: {e_img_main}")

                return {
                    "error_code": error_code_from_ai,
                    "solutions": solutions,
                    "links": links,
                    "image": image_path,
                    "is_full": bool(solutions or links),
                    "excel_row_index": index
                }

        if debug:
            pass
            # print(f"DEBUG (Excel Solution): Code '{error_code_from_ai}' not found in Excel's first column after normalization.")
        return None

    async def _find_error_solutions_internal(self, model: Optional[str] = None, debug: bool = False) -> List[Dict]:
        results = []
        model_column = None
        target_product = model or self.log_dict.get("product")

        if self.sheet and target_product:
            product_from_log = target_product
            product_normalized = product_from_log.lower().replace(" ", "")
            for cell_obj in self.sheet[2]:
                if isinstance(cell_obj.value, str):
                    cell_value_as_string = cell_obj.value
                    cell_value_normalized = cell_value_as_string.lower().replace(" ", "")
                    if cell_value_normalized == product_normalized:
                        model_column = cell_obj.column
                        break

        panic_string_original = self.log_dict.get("panicString", "")
        if not panic_string_original:
            if debug:
                pass
                # print("DEBUG: No panic string found in log_dict.")
            return results

        # Truncate panic_string_original at "slide"
        parts = panic_string_original.split("slide", 1)
        extracted_text_for_ai = parts[0].strip()

        if debug:
            pass
            # print(f"--- DEBUG START (AI Analyzer): _find_error_solutions_internal ---")
            # print(f"Original panic string: {panic_string_original[:300]}...") 
            # print(f"Text for AI (pre-slide): '{extracted_text_for_ai[:300]}...' (length: {len(extracted_text_for_ai)})")
            # print(f"Model: {target_product}, Determined Model column (for Excel): {model_column}")
        
        # Если extracted_text_for_ai пуст, _get_error_code_via_ai вернет None,
        # и последующий блок 'if not ai_determined_code:' обработает это.
        ai_determined_code = await self._get_error_code_via_ai(extracted_text_for_ai, debug=debug)

        if not ai_determined_code:
            if debug:
                pass
                # print("DEBUG: AI did not determine an error code or it was invalid. No solution will be sought from Excel.")
            return results

        if model_column is None:
            if debug:
                pass
                # print(f"DEBUG: AI determined code '{ai_determined_code}', but model column was not found. Cannot fetch Excel solutions for this model.")
            results.append({
                "error_code": ai_determined_code,
                "solutions": [],
                "links": [],
                "image": None,
                "is_full": False
            })
        else:
            solution_details = self._get_solutions_for_excel_code(ai_determined_code, model_column, debug=debug)
            if solution_details:
                results.append(solution_details)
                if debug:
                    pass
                    # if solution_details["solutions"] or solution_details["links"]:
                    #     # print(f"DEBUG: ✅ Found solution via AI+Excel for: '{solution_details['error_code']}'")
                    # else:
                    #     # print(f"DEBUG: ✅ Found error_code via AI: '{solution_details['error_code']}' but no specific solutions/links in Excel for this model column.")

                if solution_details["solutions"] or solution_details["links"] or solution_details["error_code"]:
                    mini_error_code_target = ai_determined_code + " mini"
                    if debug:
                        pass
                        # print(f"DEBUG: Ищу мини-версию с кодом '{mini_error_code_target}'")
                    if self.sheet:
                        for mini_idx, mini_row_tuple in enumerate(self.sheet.iter_rows(min_row=3, values_only=True), start=3):
                            if not mini_row_tuple or not mini_row_tuple[0]:
                                continue
                            current_mini_code_excel = str(mini_row_tuple[0]).replace('"', '').strip()
                            
                            # Улучшенная нормализация кодов для сравнения
                            excel_code_normalized = current_mini_code_excel.replace('\\/', '/').lower().strip()
                            target_code_normalized = mini_error_code_target.replace('\\/', '/').lower().strip()
                            
                            if debug:
                                pass
                                # print(f"DEBUG: Сравниваю Excel код '{current_mini_code_excel}' (нормализ.: '{excel_code_normalized}') с целевым '{mini_error_code_target}' (нормализ.: '{target_code_normalized}')")
                                # ДОПОЛНИТЕЛЬНЫЙ ЛОГ ДЛЯ ГЛУБОКОЙ ДИАГНОСТИКИ:
                                # print(f"DEBUG CMP DETAIL: Excel Original Raw: '{str(mini_row_tuple[0])}'") # Сырое значение из кортежа
                                # print(f"DEBUG CMP DETAIL: Excel Before Norm: '{current_mini_code_excel}' (len: {len(current_mini_code_excel)})")
                                # print(f"DEBUG CMP DETAIL: Excel Normalized: '{excel_code_normalized}' (len: {len(excel_code_normalized)})")
                                # print(f"DEBUG CMP DETAIL: Target AI Det. Code: '{ai_determined_code}'")
                                # print(f"DEBUG CMP DETAIL: Target Mini Code: '{mini_error_code_target}' (len: {len(mini_error_code_target)})")
                                # print(f"DEBUG CMP DETAIL: Target Normalized: '{target_code_normalized}' (len: {len(target_code_normalized)})")
                                # try:
                                #     # print(f"DEBUG CMP BYTES: Excel Norm Bytes: {excel_code_normalized.encode('utf-8', 'surrogateescape')}")
                                #     # print(f"DEBUG CMP BYTES: Target Norm Bytes: {target_code_normalized.encode('utf-8', 'surrogateescape')}")
                                # except Exception as e_encode:
                                #     # print(f"DEBUG CMP BYTES: Error encoding to bytes: {e_encode}")


                            if excel_code_normalized == target_code_normalized:
                                if debug:
                                    pass
                                    # print(f"DEBUG: Found 'mini' version in Excel: '{current_mini_code_excel}' at row ~{mini_idx}")

                                mini_solutions, mini_links = [], []
                                mini_image_path = None

                                if 0 <= (model_column - 1) < len(mini_row_tuple) and mini_row_tuple[model_column - 1]:
                                    mini_solutions, mini_links = filter_cell(str(mini_row_tuple[model_column - 1]))
                                    if debug:
                                        pass
                                        # print(f"DEBUG: Извлечено мини-решение: solutions={mini_solutions}, links={mini_links}")

                                if mini_solutions or mini_links:
                                    try:
                                        cell_mini_image = f'{get_column_letter(model_column)}{mini_idx}'
                                        mini_image = self.get_image(cell_mini_image)
                                        path_mini = f'./{self.username or "user"}_{cell_mini_image}_mini.png'
                                        mini_image.save(path_mini)
                                        mini_image_path = path_mini
                                    except ValueError:
                                        if debug:
                                            pass
                                            # print(f"DEBUG (Mini Image): No image at {cell_mini_image}")
                                    except Exception as e_mini_img:
                                        if debug:
                                            pass
                                            # print(f"DEBUG (Mini Image): Error processing image {cell_mini_image}: {e_mini_img}")

                                    mini_result_data = {
                                        "error_code": ai_determined_code,
                                        "solutions": mini_solutions,
                                        "links": mini_links,
                                        "image": mini_image_path,
                                        "is_full": False
                                    }
                                    results.append(mini_result_data)
                                    if debug:
                                        pass
                                        # print(f"DEBUG: Added 'mini' solution for '{ai_determined_code}'")
                                break
            elif ai_determined_code:
                if debug:
                    pass
                    # print(f"DEBUG: AI determined code '{ai_determined_code}', but no corresponding entry/solution found in Excel for model column {model_column}.")
                results.append({
                    "error_code": ai_determined_code,
                    "solutions": [],
                    "links": [],
                    "image": None,
                    "is_full": False
                })

        if debug:
            pass
            # print(f"--- DEBUG END (AI Analyzer): _find_error_solutions_internal ---")
        return results

    async def find_error_solutions(self, model: Optional[str] = None, debug: bool = False) -> SolutionAboutError:
        results_from_ai = await self._find_error_solutions_internal(model=model, debug=debug)

        determined_error_code = None
        
        mini_sols = []
        mini_lnks = []
        full_sols = []
        full_lnks = []

        # Извлекаем panic_string и текст для администратора до обработки результатов
        panic_str = self.log_dict.get("panicString", "")
        extracted_admin_text = ""
        if panic_str:
            parts = panic_str.split("slide", 1)
            extracted_admin_text = parts[0].strip()

        if not results_from_ai:
            return SolutionAboutError(
                descriptions=[],
                links=[],
                date_of_failure=self.log_dict.get('date', ''),
                is_full=False,
                error_code=None,
                panic_string=panic_str,
                extracted_error_text_for_admin=extracted_admin_text,
                is_mini_response_shown=False,
                has_full_solution_available=False
            )

        # Определяем базовый код ошибки. Обычно он есть в первом элементе.
        if results_from_ai[0].get('error_code'):
            determined_error_code = results_from_ai[0].get('error_code')

        for res_item in results_from_ai:
            # Если determined_error_code еще не установлен, пытаемся установить его из текущего элемента
            if not determined_error_code and res_item.get('error_code'):
                 determined_error_code = res_item.get('error_code')

            item_solutions = res_item.get('solutions', [])
            item_links = res_item.get('links', [])

            # 'is_full' is False для "мини" решений, True или отсутствует для "полных"
            if res_item.get('is_full') is False: 
                mini_sols.extend(item_solutions)
                mini_lnks.extend(item_links)
            else: 
                full_sols.extend(item_solutions)
                full_lnks.extend(item_links)
        
        user_facing_descriptions = []
        user_facing_links = []
        is_mini_shown = False
        has_full_available = False
        final_full_descriptions_for_storage = None
        final_full_links_for_storage = None
        
        # Логика приоритета "мини" ответов
        if debug:
            pass
            # print(f"DEBUG: После обработки всех результатов: mini_sols={mini_sols}, mini_lnks={mini_lnks}, full_sols={full_sols}, full_lnks={full_lnks}")
        
        if mini_sols or mini_lnks: # Если есть контент в "мини" решении
            user_facing_descriptions.extend(mini_sols)
            user_facing_links.extend(mini_lnks)
            is_mini_shown = True
            if full_sols or full_lnks: # И если также есть контент в "полном" решении
                has_full_available = True
                final_full_descriptions_for_storage = full_sols
                final_full_links_for_storage = full_lnks
                if debug:
                    pass
                    # print(f"DEBUG: Показываю МИНИ-версию с кнопкой ПОЛНОЙ версии. is_mini_shown={is_mini_shown}, has_full_available={has_full_available}")
            else:
                if debug:
                    pass
                    # print(f"DEBUG: Показываю только МИНИ-версию (нет полной). is_mini_shown={is_mini_shown}, has_full_available={has_full_available}")
        elif full_sols or full_lnks: # Если "мини" контента нет, но есть "полный"
            user_facing_descriptions.extend(full_sols)
            user_facing_links.extend(full_lnks)
            # is_mini_shown остается False
            # has_full_available остается False
            if debug:
                pass
                # print(f"DEBUG: Показываю ПОЛНУЮ версию (мини-версия не найдена). is_mini_shown={is_mini_shown}, has_full_available={has_full_available}")
        else:
            if debug:
                pass
                # print(f"DEBUG: Нет ни мини, ни полной версии. is_mini_shown={is_mini_shown}, has_full_available={has_full_available}")

        # Старый флаг is_full теперь отражает, есть ли вообще какой-то контент для пользователя
        legacy_is_full_flag = bool(user_facing_descriptions or user_facing_links)

        return SolutionAboutError(
            descriptions=user_facing_descriptions,
            links=user_facing_links,
            date_of_failure=self.log_dict.get('date', ''),
            is_full=legacy_is_full_flag, 
            error_code=determined_error_code,
            panic_string=panic_str,
            extracted_error_text_for_admin=extracted_admin_text,
            is_mini_response_shown=is_mini_shown,
            has_full_solution_available=has_full_available,
            full_descriptions=final_full_descriptions_for_storage,
            full_links=final_full_links_for_storage
        )