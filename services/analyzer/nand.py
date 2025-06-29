import io
import json
import re
import string
from typing import Any
import logging

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from pytesseract import pytesseract
from PIL import Image
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class NandList:
    def __init__(self, excel_path="./data/nand_list.xlsx"):
        self.excel_path = excel_path
        self.sheet = None
        self.header = {}
        try:
            workbook: Workbook = openpyxl.load_workbook(self.excel_path)
            self.sheet = workbook.active
            for idx, cell in enumerate(self.sheet[1]):
                if cell.value:
                    self.header[str(cell.value).strip().lower()] = idx + 1
            logger.info(f"Загружен файл NAND: {self.excel_path}. Заголовки: {self.header}")
            if not self.sheet:
                logger.error(f"Не удалось загрузить лист из файла {self.excel_path}")
            if "model" not in self.header:
                logger.error("Колонка 'model' не найдена в заголовках Excel файла.")
        except FileNotFoundError:
            logger.error(f"Файл {self.excel_path} не найден.")
        except Exception as e:
            logger.error(f"Ошибка при загрузке NAND листа из {self.excel_path}: {e}", exc_info=True)

    def find_info(self, model_name_to_find: str, lang: str):
        if not self.sheet or not self.header:
            logger.warning("NAND лист не загружен или отсутствуют заголовки, поиск невозможен.")
            return None

        model_col_idx = self.header.get("model")
        lang_col_idx = self.header.get(lang.lower())

        if not model_col_idx:
            logger.error("Индекс колонки 'model' не найден в заголовках.")
            return None
        if not lang_col_idx:
            logger.warning(f"Колонка для языка '{lang}' не найдена в заголовках. Доступные заголовки: {list(self.header.keys())}")
            if 'en' in self.header:
                lang_col_idx = self.header['en']
                logger.info("Используется колонка 'en' по умолчанию.")
            elif 'ru' in self.header:
                 lang_col_idx = self.header['ru']
                 logger.info("Используется колонка 'ru' по умолчанию (английской нет).")
            else:
                potential_lang_cols = [k for k,v in self.header.items() if k != "model"]
                if potential_lang_cols:
                    lang_col_idx = self.header[potential_lang_cols[0]]
                    logger.info(f"Используется первая доступная колонка '{potential_lang_cols[0]}' как языковая.")
                else:
                    logger.error(f"Не найдена колонка для языка '{lang}' и нет колонок по умолчанию ('en', 'ru') или других языковых колонок.")
                    return None

        logger.debug(f"Поиск информации для модели '{model_name_to_find}' на языке '{lang}'. Model col: {model_col_idx}, Lang col: {lang_col_idx}")

        for row_idx in range(2, self.sheet.max_row + 1):
            model_cell_value = self.sheet.cell(row=row_idx, column=model_col_idx).value
            if model_cell_value and str(model_cell_value).strip() == model_name_to_find:
                info = self.sheet.cell(row=row_idx, column=lang_col_idx).value
                logger.debug(f"Найдена информация для '{model_name_to_find}' на языке '{lang}': {info}")
                return info
        
        logger.warning(f"Информация для модели '{model_name_to_find}' на языке '{lang}' не найдена.")
        return None

    def get_models(self) -> list:
        if not self.sheet or "model" not in self.header:
            logger.warning("NAND лист не загружен или отсутствует колонка 'model', невозможно получить список моделей.")
            return []
        
        model_col_idx = self.header["model"]
        models = []
        for row in range(2, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row=row, column=model_col_idx).value
            if cell_value:
                models.append({
                    'name': str(cell_value).strip(),
                    'row': row
                })
        logger.debug(f"Получен список моделей (первые 5): {models[:5]}")
        return models

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    try:
        wb_test = openpyxl.load_workbook("./data/nand_list.xlsx")
        logger.info("Тестовый файл ./data/nand_list.xlsx уже существует.")
    except FileNotFoundError:
        wb_test = openpyxl.Workbook()
        sheet_test = wb_test.active
        sheet_test['A1'] = "Model"
        sheet_test['B1'] = "ru"
        sheet_test['C1'] = "en"
        sheet_test['A2'] = "H2DTDG8UD1MY"
        sheet_test['B2'] = "Информация о флеш-памяти (RU):\n- Модель: H2DTDG8UD1MY\n- Бренд: Hynix\n- Объем памяти: 16Gbyte\n- Тип памяти: BGA52\n- Поддерживаемые модели устройств:\niPhone 4, 4S\niPad 2, 3, 4"
        sheet_test['C2'] = "Information about Flash memory (EN):\n- Model: H2DTDG8UD1MY\n- Brand: Hynix\n- Memory capacity: 16Gbyte\n- Memory type: BGA52\n- Supported device models:\niPhone 4, 4S\niPad 2, 3, 4"
        sheet_test['A3'] = "TX52G6"
        sheet_test['B3'] = "Инфо для TX52G6 (RU)"
        sheet_test['C3'] = "Info for TX52G6 (EN)"
        wb_test.save("./data/nand_list.xlsx")
        logger.info("Создан тестовый файл ./data/nand_list.xlsx")

    nand_list_instance = NandList()
    
    print("\n--- Тест get_models ---")
    all_models = nand_list_instance.get_models()
    if all_models:
        print(f"Всего моделей: {len(all_models)}")
        print(f"Первые 2 модели: {all_models[:2]}")
    else:
        print("Модели не найдены.")

    print("\n--- Тест find_info ---")
    model_to_search = "H2DTDG8UD1MY"
    info_ru = nand_list_instance.find_info(model_to_search, "ru")
    info_en = nand_list_instance.find_info(model_to_search, "en")
    info_es = nand_list_instance.find_info(model_to_search, "es")

    print(f"Информация для '{model_to_search}' на RU: {info_ru}")
    print(f"Информация для '{model_to_search}' на EN: {info_en}")
    print(f"Информация для '{model_to_search}' на ES (должно быть None или по умолчанию): {info_es}")

    model_to_search_2 = "TX52G6"
    info_ru_2 = nand_list_instance.find_info(model_to_search_2, "ru")
    print(f"Информация для '{model_to_search_2}' на RU: {info_ru_2}")

    non_existent_model = "НЕ_СУЩЕСТВУЕТ"
    info_non_existent = nand_list_instance.find_info(non_existent_model, "ru")
    print(f"Информация для '{non_existent_model}' на RU: {info_non_existent}")